from datetime import datetime
import sys
import os 
import time
import openpyxl
#import xlrd

NULL_VALUE = "null"
insertCnt = 0 
isFirstLine = True
try:
	print("[FormatFile.py] 開始執行時間：" + time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))

	if len(sys.argv) < 3 :
		print("You need input two parameter(fmt : 股票代號 資料抓取日期)")
		print("syntax : C:\python FormatStocksSaleMonData.py 2002 20220420")
		sys.exit()

	loadFileDir = "Data\\EXCEL\\Tranfer\\salemon\\"
#	loadFileDir = "C:\Users\linea\OneDrive\myStocksPGMs\V2.0\xls2xlsx\transfer_ok\saleMonth\\"
	saveFileDir = "Data\\TXT\\salemon\\"
#	inputFile = sys.argv[1] + ".xlsx"
	stockCode = sys.argv[1]
	inputFile = stockCode + "-salemon-" + sys.argv[2] +".xlsx"
	outputFile = stockCode + ".txt"
	
	outfile = open(saveFileDir + outputFile, 'w')

#	wb = openpyxl.load_workbook('csv\\goodinfo\\saleMonth\\' + inputFile)
	print(loadFileDir + inputFile)
	wb = openpyxl.load_workbook(loadFileDir + inputFile)
	sheet = wb.worksheets[0]

	isSTOP = False
	irow = 5
	icol = 1

#	theList.append("insert into stocks_sale_month (stock_no,date,open_price_mon,end_price_mon,hgh_price_mon,low_price_mon,ups_downs,ups_downs_p,mon_revenue,mon_increase_in_revenue,ann_ins_revenue,cum_revenue,ann_ins_cum_revenue_p,csd_revenue,mon_ins_csd_revenue_p,ann_ins_csd_revenue_p,cum_csd_revenue,ann_ins_cum_csd_revenue_p) values ('" + stockCode + "'")
#	theSQLCmd = "insert into stocks_sale_month (stock_no, date, open_price_mon, end_price_mon, hgh_price_mon, low_price_mon, ups_downs, ups_downs_p, mon_revenue, mon_increase_in_revenue, ann_ins_revenue, cum_revenue, ann_ins_cum_revenue_p, csd_revenue, mon_ins_csd_revenue_p, ann_ins_csd_revenue_p, cum_csd_revenue, ann_ins_cum_csd_revenue_p) values ('%s', '%d', '%f',  '%f',  '%f',  '%f',  '%f',  '%f',  '%f', '%f', '%f', '%f', '%f', '%f', '%f', '%f', '%f', '%f')"
#	outfile.write(theSQLCmd+"\n")
	while not isSTOP :
		theList = []
		theList.append("insert into stocks_sale_month (stock_no,date,open_price_mon,end_price_mon,hgh_price_mon,low_price_mon,ups_downs,ups_downs_p,mon_revenue,mon_increase_in_revenue,ann_ins_revenue,cum_revenue,ann_ins_cum_revenue_p,csd_revenue,mon_ins_csd_revenue_p,ann_ins_csd_revenue_p,cum_csd_revenue,ann_ins_cum_csd_revenue_p) values ('" + stockCode + "'")

		for icol in range(1, 18) :
			theValue = sheet.cell(row = irow, column = icol).value
#			print(theValue)
#			if sheet.cell(row = irow, column = icol).value == None :
			if icol == 1 and sheet.cell(row = irow, column = icol).value == None :
					isSTOP = True
					theList.pop()
					break
			elif type(theValue) == str :
				if theValue == "-" :
					theValue = NULL_VALUE
				theList.append(theValue)
			elif type(theValue) == int :
				theList.append(theValue)
			elif type(theValue) == float :
				theList.append(theValue)
			elif type(theValue) == datetime :
				theList.append(str(theValue.strftime("%Y%m%d")))
			else :
				theList.append(theValue)
		print(theList)
		if len(theList) > 0 :
			outfile.write(", ".join([str(_) for _ in theList]))
			outfile.write(");\n")
		insertCnt += 1

		# 預防錯誤，理論上應該不會超過100列
		if irow > 100 :
			print("i=" + str(irow))
			isSTOP = True
		irow += 1

	outfile.close()
	
except IOError as err :
	print('File error : ' + str(err))

print("資料處理完成!! 共 " + str(insertCnt) + " 筆。")
print("[FormatFile.py] 結束執行時間：" + time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))

"""
stock_no
dividend_year
cash_div_surplus
cash_div_reserve
total_cash_div
stock_div_surplus
stock_div_reserve
total_stock_div
total_dividend
total_div_cash
total_div_stocks
days_fill_cash
days_fill_stocks
stock_price_year
year_high_price
year_low_price
year_avg_price
avg_ann_cash_yield
avg_ann_stock_yield
avg_ann_yield
period_of_dividend
eps
div_earnings_dis_ratio
alo_earnings_dis_ratio
earnings_dis_ratio
"""